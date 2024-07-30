# Import the necessary libraries
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Font, Alignment
from openpyxl.cell import Cell
from openpyxl import Workbook

corner_borders = Border(left=Side(style='thick'), 
                     right=Side(style='thick'))

top_border = Border(top=Side(style='thick'),
                    left=Side(style='thick'), 
                     right=Side(style='thick'))

bottom_border = Border(bottom=Side(style='thick'),
                       left=Side(style='thick'), 
                     right=Side(style='thick'))

first_row = 1
last_row = first_row + 6
first_collumn = 1

wb = Workbook()
ws = wb.active
ws.column_dimensions['A'].width = 27
ws.column_dimensions['C'].width = 27
ws.column_dimensions['E'].width = 27
ws.column_dimensions['H'].width = 27

for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=first_collumn, max_col=first_collumn):
    for cell in row:
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(name = 'Arial')
        if cell.row == first_row:
            cell.value = '++++'
            cell.border = top_border
        elif cell.row == last_row:
            cell.border = bottom_border
        else:
            cell.border = corner_borders

        if cell.row == last_row - 2:
            cell.fill = PatternFill(start_color="0000B050", fill_type = "solid")
            cell.font = Font(color="FFFFFF", size=12, bold=True)
            cell.value = 'LOJA'
# Save the workbook to a file
wb.save("Jit.xlsx")
# Print a success message
print("Excel file created successfully!")