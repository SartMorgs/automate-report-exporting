from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from report_feeding.report.jit.utils import Utils as ReportFeedingUtils


class LayoutBuilder:
    """
    Report Feeding: Layout Builder Class
    
    Used for set the layout for each jit card.
    
    Attributes:
        filename (str): Report filename to be saved.
    """
    def __init__(self, filename: str):
        self.__CORNER_BORDER = Border(left=Side(style=ReportFeedingUtils.BORDER_STYLE), right=Side(style=ReportFeedingUtils.BORDER_STYLE))
        self.__TOP_BORDER = Border(top=Side(style=ReportFeedingUtils.BORDER_STYLE), left=Side(style=ReportFeedingUtils.BORDER_STYLE), right=Side(style=ReportFeedingUtils.BORDER_STYLE))
        self.__BOTTOM_BORDER = Border(bottom=Side(style=ReportFeedingUtils.BORDER_STYLE), left=Side(style=ReportFeedingUtils.BORDER_STYLE),  right=Side(style=ReportFeedingUtils.BORDER_STYLE))
        
        # Layout Details
        self.__HORIZONTAL_ALIGNMENT = "center"
        self.__FIRST_ROW_VALUE = "\'++++"
        self.__FONT = "Arial"
        
        self.__LAYOUT_CONFIG = {
            ReportFeedingUtils.STORE_VALUE: {
                ReportFeedingUtils.BACKGROUND_COLOR_KEY: ReportFeedingUtils.STORE_BACKGROUND_COLOR,
                ReportFeedingUtils.FILL_TYPE_KEY: ReportFeedingUtils.FILL_TYPE,
                ReportFeedingUtils.FONT_COLOR_KEY: ReportFeedingUtils.FONT_COLLOR,
                ReportFeedingUtils.FONT_SIZE_KEY: ReportFeedingUtils.FONT_SIZE,
                ReportFeedingUtils.VALUE_KEY: ReportFeedingUtils.STORE_VALUE
            },
            ReportFeedingUtils.REPRO_VALUE: {
                ReportFeedingUtils.BACKGROUND_COLOR_KEY: ReportFeedingUtils.REPRO_BACKGROUND_COLOR,
                ReportFeedingUtils.FILL_TYPE_KEY: ReportFeedingUtils.FILL_TYPE,
                ReportFeedingUtils.FONT_COLOR_KEY: ReportFeedingUtils.FONT_COLLOR,
                ReportFeedingUtils.FONT_SIZE_KEY: ReportFeedingUtils.FONT_SIZE,
                ReportFeedingUtils.VALUE_KEY: ReportFeedingUtils.REPRO_VALUE
            }
        }
        
        self.filename = filename

    def __get_border_type_dict(self, first_row_index: int, last_row_index: int):
        """
        Gets the dictinary with the border config for each specific row type.
        
        Args:
            first_row_index (int): First row index to be used in the jit card.
            last_row_index (int): Last row index to be used in the jit card.
        
        Returns:
            The dictinary with the border config for each specific row type. (Dict[str, str])
        """
        return {
            first_row_index: self.__TOP_BORDER,
            last_row_index: self.__BOTTOM_BORDER
        }

    def __set_row_standard_layout_for_jit_card(self, cell: Cell, jit_card_type: str):
        """
        Sets row standard layout for each jit card. It sets the color, fill type, font and value as ```store``` or ```repro```.
        
        Args:
            cell (Cell): Cell in the spreadsheet to change.
            jit_card_type (str): The jit card type, ```store``` or ```repro```.
        """
        cell.fill = PatternFill(
            start_color=self.__LAYOUT_CONFIG[jit_card_type][ReportFeedingUtils.BACKGROUND_COLOR_KEY],
            fill_type=self.__LAYOUT_CONFIG[jit_card_type][ReportFeedingUtils.FILL_TYPE_KEY]
        )
        cell.font = Font(
            color=self.__LAYOUT_CONFIG[jit_card_type][ReportFeedingUtils.FONT_COLOR_KEY],
            size=self.__LAYOUT_CONFIG[jit_card_type][ReportFeedingUtils.FONT_SIZE_KEY], bold=True
        )
        cell.value = self.__LAYOUT_CONFIG[jit_card_type][ReportFeedingUtils.VALUE_KEY]

    def __set_row_layout(self, cell: Cell, first_row_index: int, last_row_index: int, jit_card_type: str):
        """
        Sets row layout: alignment, font, border, font, color, etc...
        
        Args:
            cell (Cell): Cell in the spreadsheet to change.
            first_row_index (int): First row index to be used in the jit card.
            last_row_index (int): Last row index to be used in the jit card.
            jit_card_type (str): The jit card type, ```store``` or ```repro```.
        """
        cell.alignment = Alignment(horizontal=self.__HORIZONTAL_ALIGNMENT)
        cell.font = Font(name=self.__FONT)
        
        border_layout = self.__get_border_type_dict(first_row_index, last_row_index)
        cell.border = border_layout[cell.row] if cell.row in border_layout else self.__CORNER_BORDER
        if cell.row == first_row_index:
            cell.value = self.__FIRST_ROW_VALUE

        if cell.row == last_row_index - 2:
            self.__set_row_standard_layout_for_jit_card(cell, jit_card_type)

    def _generate_layout(
        self, active_ws: Workbook, first_row_index: int, last_row_index: int, column_index: int, jit_card_type: str):
        """
        Generates layout for each jit card.
        
        Args:
            active_ws (Workbook): The active worksheet.
            first_row_index (int): The first row of the jit card.
            last_row_index (int): The last row of the jit card.
            column_index (str): The column index.
            jit_card_type (str): The jit card type, ```store``` or ```repro```.
        """
        for row in active_ws.iter_rows(min_row=first_row_index, max_row=last_row_index, min_col=column_index, max_col=column_index):
            for cell in row:
                self.__set_row_layout(cell, first_row_index, last_row_index, jit_card_type)
